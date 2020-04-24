import cv2
import numpy as np
from threading import Thread
import time
import datetime
from os import path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import matplotlib.pyplot as plt
import pytesseract 

pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files/Tesseract-OCR/tesseract.exe'
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
cap = cv2.VideoCapture(0)

counter_frames = 1
done =0
detected =0
text = ""
workbook_name = "Attendance.xlsx"
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
mouth_cascade = cv2.CascadeClassifier('haarcascade_mcs_mouth.xml')
nose_cascade = cv2.CascadeClassifier('haarcascade_mcs_nose.xml')
hand_cascade = cv2.CascadeClassifier('hand.xml')
names_dict = {"1501075":"Martin Joseph","1500935":"Omar Hesham","1500920":"Omar mohamed","1501333":"Mahmoud Ibrahim"}
out = np.zeros((640,480), np.uint8)

# To capture video from webcam. 
cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_SIMPLEX 
  
# org 
org = (50, 30) 
  
# fontScale 
fontScale = 0.6
   
# Blue color in BGR 
color_pass = (0, 255, 0)
color_stop = (0, 0, 255) 
color_orange = (52,177,235)
  
# Line thickness of 2 px 
thickness = 2
counter_pass =1
counter_pass_threshold =10

        ##########
if path.exists(workbook_name):
    wb =  load_workbook(filename = workbook_name)
    sheet = wb.active                
else:
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "ID"
    sheet["B1"] = "Time"
    sheet["C1"] = "Date"  

flag_thread = 0
def text_checking():
    global text,done,detected
    text = text.replace("-", "")
    text = text.replace(".", "")
    text = text.replace(' ', "")
#    text = text.replace('\n', ' ').replace('\r', '')
    text = text.rstrip()
    print("text length = ",len(text))
    print("ID is ",text)
    if text in names_dict :
        done =1
        detected=0
        now = datetime.datetime.now()
        cells =sheet["A"]            
        for cell in cells:
           if text ==cell.value:
               break
        else:
            new_Attendance= [text,str(now.hour)+":"+str(now.minute)+":"+str(now.second),str(now.day)+"/"+str(now.month)+"/"+str(now.year)]
            sheet.append(new_Attendance)                            
def employee_info(img):
    global out
    blk = np.zeros(img.shape, np.uint8)
    profile = cv2.imread("images/"+names_dict[text]+".jpg")
    profile = cv2.resize(profile, (180,160))
    img[60:220,450:630] = profile
    cv2.rectangle(blk, (450, 230), (630, 420), (255, 255, 255), cv2.FILLED)
    cv2.putText(img, 'ID:', (455,250), font,  0.7, (0,0,0), 2, cv2.LINE_AA)
    cv2.putText(img, text, (455,280), font,  0.7, (0,0,255), 2, cv2.LINE_AA)
    cv2.putText(img, 'Name:', (455,310), font,  0.7, (0,0,0), 2, cv2.LINE_AA)
    cv2.putText(img, names_dict[text], (455,340), font,  0.7, (0,0,255), 2, cv2.LINE_AA)
    cv2.putText(img, 'Arrival time:', (455,370), font,  0.7, (0,0,0), 2, cv2.LINE_AA)
    cv2.putText(img, str(datetime.datetime.now().hour)+":"+str(datetime.datetime.now().minute), (455,400), font,  0.7, (0,0,255), 2, cv2.LINE_AA)
    out = cv2.addWeighted(img, 1.0, blk, 0.25, 1)
#def delay_time():
#    print("hello_from_thread")
#    global flag_thread
#    time.sleep(3)
#    flag_thread = 1
#
#thread = Thread(target=delay_time, args=())

while True:
    # Read the frame
    _, img = cap.read()
    if flag_thread == 0:
        img_cpy = img.copy()
#        img= cv2.flip(img,  1);
        img = cv2.line(img, (200,0), (200,900), color_pass, thickness)
        img = cv2.line(img, (450,0), (450,900), color_pass, thickness)
        cv2.putText(img, 'You hand here', (50,450), font,  fontScale, color_orange, thickness, cv2.LINE_AA)
        cv2.putText(img, 'You face here', (250,450), font,  fontScale, color_orange, thickness, cv2.LINE_AA)
        cv2.putText(img, 'You hand here', (460,450), font,  fontScale, color_orange, thickness, cv2.LINE_AA)
        cv2.rectangle(img, (220, 70), (430, 300), (0, 255, 0), 2)
        #####gloves
        gloves_img = img_cpy.copy()
        gloves_img[:, 200:450, [0, 1, 2]] = 0
        #plt.imshow(gloves_img)
        hsv_frame = cv2.cvtColor(gloves_img, cv2.COLOR_BGR2HSV)
        # Blue color
        low_blue = np.array([94, 80, 2])
        high_blue = np.array([126, 255, 255])
        blue_mask = cv2.inRange(hsv_frame, low_blue, high_blue)
        blue = cv2.bitwise_and(gloves_img, gloves_img, mask=blue_mask)
        
        # Convert to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # Detect the faces
        faces = face_cascade.detectMultiScale(gray, 1.1, 5)
        mouth_rects = []
        nose_rects = []
        # Draw the rectangle around each face
        for (x, y, w, h) in faces:
            #cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)
            mouth_rects = mouth_cascade.detectMultiScale(gray[y+int((h/2)):y+h,x:x+w], 1.3,11)
            nose_rects = nose_cascade.detectMultiScale(gray[y+int(h/3):y+h,x:x+w], 1.3, 4)
            for (mx,my,mw,mh) in mouth_rects:
                my = int(my - 0.15*mh)
                cv2.rectangle(img, (x+mx,y+int(h/2)+my), (x+mx+mw,y+int(h/2)+my+mh), (0,0,255), 3)
                break
            for (nx,ny,nw,nh) in nose_rects:
                cv2.rectangle(img, (x+nx,y+ny+int(h/3)), (x+nx+nw,y+ny+nh+int(h/3)), (255,0,0), 3)
                break
            break
        
        white_pixels = np.count_nonzero(blue_mask == 255)
        if len(faces) == 0:
            cv2.putText(img, 'Noone detected', org, font,  fontScale, color_stop, thickness, cv2.LINE_AA)
        elif(len(faces)>len(mouth_rects) and len(faces)>len(nose_rects)) and white_pixels > 35000:
            counter_pass +=1
            print(counter_pass)
            if counter_pass >= counter_pass_threshold:
                cv2.putText(img, 'You can pass', org, font,  fontScale, color_pass, thickness, cv2.LINE_AA)
            if counter_pass == counter_pass_threshold:
                flag_thread = 1
                counter_frames = 1
                done =0
                detected =0

        elif(len(faces)>len(mouth_rects) and len(faces)==len(nose_rects)):
            cv2.putText(img, 'You are not wearing mask correctly', org, font,  fontScale, color_orange, thickness, cv2.LINE_AA)
        elif len(faces)==len(mouth_rects) and len(faces)==len(nose_rects):
            cv2.putText(img, 'You must wear a mask to pass', org, font,  fontScale, color_stop, thickness, cv2.LINE_AA)
        elif(len(faces)>len(mouth_rects) and len(faces)>len(nose_rects)):
            cv2.putText(img, 'the mask is in correct position', org, font,  fontScale, color_pass, thickness, cv2.LINE_AA)
        if white_pixels < 10000:
            cv2.putText(img, 'You should wear gloves', (50,60), font,  fontScale, color_stop, thickness, cv2.LINE_AA)
        else:
            cv2.putText(img, 'the gloves are worn', (50,60), font,  fontScale, color_pass, thickness, cv2.LINE_AA)
    elif flag_thread == 1:
        cropped_image = img[70:230,90:340]
        gray = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
        if done == 0:
            cv2.rectangle(img, (90, 70), (340, 230), (0, 0, 255), 2) #card
            cv2.putText(img, 'place your card here', (100,100), font,  0.7, (0,0,255), 1, cv2.LINE_AA)
        if detected ==1 and done == 0:
            cv2.putText(img, 'Scanning....', (120,130), font,  1, (0,255,255), 1, cv2.LINE_AA)
        elif done ==1:
            cv2.rectangle(img, (90, 70), (340, 230), (0, 255, 0), 2) #card
            cv2.putText(img, 'Scanning Done', (100,150), font,  1, (0,255,0), 2, cv2.LINE_AA)
            employee_info(img)
            flag_thread = 0
            counter_pass =1
            cv2.imshow('img', out)
            k = cv2.waitKey(3000)
    
        face_crop = gray[:,150:250]
        faces = face_cascade.detectMultiScale(face_crop, 1.1, 5)
        for (x, y, w, h) in faces:
            x += 160
            cv2.rectangle(img, (x+80, y+70), (x+80+w, y+70+h), (0, 255, 0), 2)
        if len(faces) == 1:
            detected = 1
            counter_frames+=1
            print(counter_frames)
            
    
        if len(faces) == 1 and counter_frames%30 ==0 and done ==0:
            gray_cropped = gray[y:y+h+50,x-150:x-60]
            cv2.imwrite("F:/test"+str(counter_frames)+'.jpg',gray_cropped)
            # Preprocessing the image starts 
            
            gray_cropped = cv2.resize(gray_cropped, (gray_cropped.shape[0]*3,gray_cropped.shape[1]*3))
            gray = gray_cropped
            _,thresh_ = cv2.threshold(gray,127,255,cv2.THRESH_BINARY)
            ret, thresh1 = cv2.threshold(thresh_, 0, 255, cv2.THRESH_OTSU | cv2.THRESH_BINARY_INV) 
            
            # Specify structure shape and kernel size. 
            # Kernel size increases or decreases the area 
            # of the rectangle to be detected. 
            # A smaller value like (10, 10) will detect 
            # each word instead of a sentence. 
            rect_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (18, 18)) 
            
            # Appplying dilation on the threshold image 
            dilation = cv2.dilate(thresh1, rect_kernel, iterations = 1) 
            
            # Finding contours 
            contours, hierarchy = cv2.findContours(dilation, cv2.RETR_EXTERNAL, 
            												cv2.CHAIN_APPROX_NONE) 
            
            # Creating a copy of image 
            cropped_image = cropped_image[y:y+h+50,x-150:x-60]
            cropped_image = cv2.resize(cropped_image, (cropped_image.shape[0]*3,cropped_image.shape[1]*3))
            im2 = cropped_image.copy() 
            # A text file is created and flushed 
            
            # Looping through the identified contours 
            # Then rectangular part is cropped and passed on 
            # to pytesseract for extracting text from it 
            # Extracted text is then written into the text file 
            for cnt in contours:
                x, y, w, h = cv2.boundingRect(cnt)
                rect = cv2.rectangle(im2, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cropped = im2[y:y + h, x:x + w]
                custom_config = r'--oem 3 --psm 6 outputbase digits'
                text = pytesseract.image_to_string(cropped, config=custom_config)
                if len(text) == 7:
                    break
            text_checking()

    # Display
    cv2.imshow('img', img)
    # Stop if escape key is pressed
    k = cv2.waitKey(30) & 0xff
    if k==27:
        cv2.destroyAllWindows();
        break
# Release the VideoCapture object
cap.release()
wb.save(filename=workbook_name)   
